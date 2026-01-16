import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.shortcuts import render, get_object_or_404, redirect # <--- IMPORTANTE: redirect añadido
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.db.models import Sum
from django.utils import timezone
from datetime import datetime, timedelta
from xhtml2pdf import pisa
from .models import GuiaEntrega, Producto, Gasto, Cliente
from django.contrib import admin
from django.contrib.auth.decorators import login_required

# --- VISTA 1: GENERADOR DE PDF ---
def generar_pdf_guia(request, guia_id):
    guia = get_object_or_404(GuiaEntrega, pk=guia_id)
    detalles = guia.detalles.all()
    saldo_pendiente = guia.total_venta - guia.monto_cobrado
    
    context = {
        'guia': guia,
        'detalles': detalles,
        'saldo_pendiente': saldo_pendiente,
    }
    
    template_path = 'gestion/guia_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    
    # Lógica para ver en modal (inline) vs descargar (attachment)
    if request.GET.get('ver') == 'true':
        disposition = 'inline'
    else:
        disposition = 'attachment'

    response['Content-Disposition'] = f'{disposition}; filename="Guia-{guia.numero_guia}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Tuvimos errores <pre>' + html + '</pre>')
    return response

# --- VISTA 2: DASHBOARD GERENCIAL (PROTEGIDO 🔒) ---
@login_required(login_url='/adminconfiguracion/login/')
def dashboard_analiticas(request):
    
    # ⛔ SEGURIDAD: Si el usuario NO es el jefe (superuser),
    # lo expulsamos del dashboard y lo mandamos al menú principal.
    if not request.user.is_superuser:
        return redirect('/adminconfiguracion/') 
    
    # 1. Definir fechas
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    fecha_inicio = request.GET.get('fecha_inicio', inicio_mes.strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', hoy.strftime('%Y-%m-%d'))

    # 2. Filtrar
    ventas = GuiaEntrega.objects.filter(fecha_emision__range=[fecha_inicio, fecha_fin])
    gastos = Gasto.objects.filter(fecha_emision__range=[fecha_inicio, fecha_fin])

    # 3. Calcular
    total_ventas = ventas.aggregate(total=Sum('total_venta'))['total'] or 0
    total_cobrado = ventas.aggregate(total=Sum('monto_cobrado'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    deuda_calle = 0
    for g in GuiaEntrega.objects.exclude(estado_pago='PAGADO'):
        deuda_calle += (g.total_venta - g.monto_cobrado)

    ganancia_neta = total_ventas - total_gastos

    productos_bajo_stock = Producto.objects.filter(stock_actual__lte=10).order_by('stock_actual')[:5]
    ultimas_ventas = ventas.order_by('-fecha_emision')[:10]

    # 4. PREPARAR EL MENÚ LATERAL
    context = admin.site.each_context(request)
    
    context.update({
        'total_ventas': total_ventas,
        'total_cobrado': total_cobrado,
        'total_gastos': total_gastos,
        'ganancia_neta': ganancia_neta,
        'deuda_calle': deuda_calle,
        'productos_bajo_stock': productos_bajo_stock,
        'ultimas_ventas': ultimas_ventas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })
    
    return render(request, 'gestion/dashboard.html', context)

# --- VISTA 3: EXPORTAR EXCEL (PROTEGIDO 🔒) ---
@login_required(login_url='/adminconfiguracion/login/')
def exportar_reporte_excel(request):
    
    # ⛔ SEGURIDAD: Los colaboradores tampoco pueden bajar la base de datos
    if not request.user.is_superuser:
        return redirect('/adminconfiguracion/')

    # 1. Obtener filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    fecha_inicio = request.GET.get('fecha_inicio', inicio_mes.strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', hoy.strftime('%Y-%m-%d'))

    # 2. Consultar datos
    ventas = GuiaEntrega.objects.filter(fecha_emision__range=[fecha_inicio, fecha_fin])
    gastos = Gasto.objects.filter(fecha_emision__range=[fecha_inicio, fecha_fin])

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    
    # --- HOJA 1: RESUMEN Y VENTAS ---
    ws = wb.active
    ws.title = "Reporte Ventas"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    
    ws['A1'] = f"REPORTE DE MOVIMIENTOS: {fecha_inicio} al {fecha_fin}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    headers = ["Fecha", "Guía N°", "Cliente", "Estado", "Total (S/.)"]
    ws.append([]) 
    ws.append(headers)
    
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font

    total_ventas = 0
    for v in ventas:
        ws.append([
            v.fecha_emision,
            v.numero_guia,
            v.cliente.nombre_contacto,
            v.get_estado_pago_display(),
            v.total_venta
        ])
        total_ventas += v.total_venta
    
    ws.append(["", "", "", "TOTAL VENTAS:", total_ventas])
    ws[f'E{ws.max_row}'].font = Font(bold=True)

    # --- HOJA 2: GASTOS ---
    ws2 = wb.create_sheet("Gastos")
    ws2['A1'] = "DETALLE DE GASTOS Y COMPRAS"
    ws2['A1'].font = Font(bold=True, size=14)

    headers_gastos = ["Fecha", "Proveedor", "Descripción", "Categoría", "Estado", "Monto (S/.)"]
    ws2.append([])
    ws2.append(headers_gastos)
    
    for cell in ws2[3]:
        cell.fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
        cell.font = header_font

    total_gastos = 0
    for g in gastos:
        ws2.append([
            g.fecha_emision,
            g.proveedor.razon_social,
            g.descripcion,
            g.get_categoria_display(),
            g.get_estado_display(),
            g.monto
        ])
        total_gastos += g.monto

    ws2.append(["", "", "", "", "TOTAL GASTOS:", total_gastos])
    ws2[f'F{ws2.max_row}'].font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_MyM_{fecha_inicio}_{fecha_fin}.xlsx'
    
    wb.save(response)
    return response

# --- VISTA 4: HEALTH CHECK ---
def health_check(request):
    return HttpResponse("OK")

# --- VISTA 5: API PARA PRECIOS (NECESARIA PARA ADMIN) ---
def obtener_info_producto(request, producto_id):
    from .models import Producto
    try:
        producto = Producto.objects.get(pk=producto_id)
        data = {
            'precio': producto.precio_unitario,
            'stock': producto.stock_actual
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

# --- VISTA 6: API PARA DIRECCIÓN (NECESARIA PARA ADMIN) ---
def obtener_direccion_cliente(request, cliente_id):
    from .models import Cliente
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
        return JsonResponse({'direccion': cliente.direccion_principal})
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)